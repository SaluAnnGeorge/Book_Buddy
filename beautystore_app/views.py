from datetime import date
from decimal import Decimal
from django.shortcuts import render,redirect
from django.http import HttpResponse
from .models import Product
from django.views import View
from django.shortcuts import render,get_object_or_404,redirect
from .forms import *
from django.conf import settings
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import login as auth_login
from django.contrib import messages
from .models import Wishlist
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from django.contrib.auth import login, authenticate, logout
from django.core.exceptions import ValidationError
from django.urls import reverse_lazy
from .models import Customer
from django.contrib.auth.password_validation import validate_password
from django.contrib.auth.decorators import login_required
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.contrib.sites.shortcuts import get_current_site
from django.core.mail import EmailMessage
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from .models import *
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from .forms import AddressForm
from django.template.loader import get_template
# from xhtml2pdf import pisa
# Create your views here.

# return HttpResponse("<h1>Welcome to django</h1>")
   
def about (request):
    return HttpResponse("<h1>About page</h1>")

def index(request):
    products = Product.objects.all()
    categories = Category.objects.prefetch_related('subcategory_set').all()
    return render(request, 'index.html', {'products': products, "categories":categories})


def product_add(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('admin_product_page')  # Redirect to a success page after adding the product
    else:
        form = ProductForm()
    return render(request, 'add_product.html', {'form': form})
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductUpdateForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            return redirect('admin_product_page')  # Redirect to the product detail page after updating
    else:
        form = ProductUpdateForm(instance=product)
    return render(request, 'update_product.html', {'form': form, 'product': product})
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        return redirect('admin_product_page')  # Redirect to the product list page after deletion
    return render(request, 'delete_product.html', {'product': product})
def admin_product_page(request):
    products = Product.objects.all()
    return render(request, 'admin_product_page.html', {'products': products})
def admin_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)

            if user is not None and user.is_superuser:
                auth_login(request, user)
                return redirect('admin_product_page')
            else:
                return render(request, 'admin_login.html', {'form': form, 'error_message': 'Invalid credentials'})

    else:
        form = AuthenticationForm()

    return render(request, 'admin_login.html', {'form': form})




def product_search(request):
    query = request.GET.get('q')
    products = Product.objects.filter(name__icontains=query)
    return render(request, 'search_results.html', {'products': products})

def product_details(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    return render(request, 'product_details.html', {'product': product})

@login_required(login_url='login')
def wishlist(request):
    user = request.user
    wishlist = Wishlist.objects.filter(user=user)
    return render(request, 'wishlist.html', {'wishlist': wishlist})

@login_required(login_url='login')
def add_to_wishlist(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    wishlist, created = Wishlist.objects.get_or_create(user=request.user, product=product)
    return redirect('wishlist')

def remove_from_wishlist(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    Wishlist.objects.filter(user=request.user, product=product).delete()
    return redirect('wishlist')

def customer_dashboard(request):
    return render(request, 'customer_dashboard.html')

def register(request):
    if request.method == 'POST':
        full_name = request.POST.get('full-name')
        email = request.POST.get('your-email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm-password')
        phone_number = request.POST.get('phone-number')

        # Check if passwords match
        if password != confirm_password:
            messages.error(request, 'Passwords Do Not Match!')
            return render(request, 'register.html')

        # Check password strength using Django's built-in validators
        try:
            validate_password(password)
        except ValidationError as e:
            messages.error(request, ', '.join(e.messages))
            return render(request, 'register.html')

        # Create user
        try:
            user = User.objects.create_user(username=email, email=email, password=password)
        except:
            messages.error(request, 'Email already exists!')
            return render(request, 'register.html')

        # Create customer
        customer = Customer.objects.create(user=user, customer_name=full_name, email=email, contact_number=phone_number)
        
        # Authenticate and login user
        user = authenticate(request, username=email, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, 'Your Account Has Been Registered Successfully!')
            return redirect('index')
        else:
            messages.error(request, 'Failed to login user.')
            return render(request, 'register.html')

    return render(request, 'register.html')

def login_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        user = authenticate(request, username=email, password=password)
        
        if user is not None:
            try:
                login(request, user)
                messages.success(request, 'You have successfully logged in!')
                return redirect('index')  # Change 'index' to your desired redirect URL after login
            except Exception as e:
                # If there's an exception during login, print the exception for debugging
                print(e)
                messages.error(request, 'An error occurred during login. Please try again.')
                return render(request, 'login.html')  # Change 'login.html' to your login template path
        else:
            messages.error(request, 'Invalid email or password. Please try again.')
            return render(request, 'login.html')  # Change 'login.html' to your login template path
    return render(request, 'login.html')  # Change 'login.html' to your login template path

def user_logout(request):
    logout(request)
    return redirect('index') 



def change_password(request):
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        user = request.user  # Assuming the user is already logged in

        if user.check_password(old_password):
            if new_password == confirm_password:
                user.set_password(new_password)
                user.save()
                messages.success(request, "Password changed successfully.")
                return redirect('login')
            else:
                messages.error(request, "New passwords do not match.")
        else:
            messages.error(request, "Old password is incorrect.")

    return render(request, 'change_password.html')

class CustomTokenGenerator(PasswordResetTokenGenerator):
    def _make_hash_value(self, user, timestamp):
        return (
            str(user.pk) + user.password + str(timestamp)
        )



def forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            token = CustomTokenGenerator().make_token(user)
            reset_password_url = request.build_absolute_uri('/reset_password/{}/{}/'.format(uid, token))
            email_subject = 'Reset Your Password'

            # Render both HTML and plain text versions of the email
            email_body_html = render_to_string('reset_password_email.html', {
                'reset_password_url': reset_password_url,
                'user': user,
            })
            email_body_text = "Click the following link to reset your password: {}".format(reset_password_url)

            # Create an EmailMultiAlternatives object to send both HTML and plain text versions
            email = EmailMultiAlternatives(
                email_subject,
                email_body_text,
                settings.EMAIL_HOST_USER,
                [email],
            )
            email.attach_alternative(email_body_html, 'text/html')  # Attach HTML version
            email.send(fail_silently=False)

            messages.success(request, 'An email has been sent to your email address with instructions on how to reset your password.')
            return redirect('login')
        except User.DoesNotExist:
            messages.error(request, "User with this email does not exist.")
    return render(request, 'forgot_password.html')


def reset_password(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and CustomTokenGenerator().check_token(user, token):
        if request.method == 'POST':
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')

            if new_password == confirm_password:
                user.set_password(new_password)
                user.save()
                messages.success(request, "Password reset successfully. You can now login with your new password.")
                return redirect('login')
            else:
                messages.error(request, "Passwords do not match.")
        return render(request, 'reset_password.html')
    else:
        messages.error(request, "Invalid reset link. Please try again or request a new reset link.")
        return redirect('login')

@login_required
def edit_customer(request):
    # Retrieve the current logged-in user
    current_user = request.user
    # Check if the current user has a corresponding Customer instance
    try:
        customer = Customer.objects.get(user=current_user)
    except Customer.DoesNotExist:
        # Handle the case where the logged-in user does not have a corresponding Customer instance
        return HttpResponse("You are not associated with any customer profile.")

    if request.method == 'POST':
        # Update customer details with the data from the form
        customer.customer_name = request.POST['full-name']
        customer.email = request.POST['your-email']
        customer.contact_number = request.POST['phone-number']
        customer.save()

        # Update associated user's email
        current_user.email = request.POST['your-email']
        current_user.username = request.POST['your-email']
        current_user.save()

        # Redirect to the customer detail page after editing
        return redirect('index')

    # If it's a GET request, display the edit form with existing customer details
    return render(request, 'edit_customer.html', {'customer': customer})

@login_required
def address_list(request):
    addresses = Address.objects.filter(customer=request.user.customer)
    return render(request, 'address_list.html', {'addresses': addresses})

@login_required
def address_create(request):
    if request.method == 'POST':
        form = AddressForm(request.POST)
        if form.is_valid():
            address = form.save(commit=False)
            address.customer = request.user.customer
            address.save()
            return redirect('address_list')
    else:
        form = AddressForm()
    return render(request, 'address_form.html', {'form': form})

@login_required
def address_edit(request, pk):
    address = get_object_or_404(Address, pk=pk, customer=request.user.customer)
    if request.method == 'POST':
        form = AddressForm(request.POST, instance=address)
        if form.is_valid():
            form.save()
            return redirect('address_list')
    else:
        form = AddressForm(instance=address)
    return render(request, 'address_form.html', {'form': form})

@login_required
def address_delete(request, pk):
    address = get_object_or_404(Address, pk=pk, customer=request.user.customer)
    if request.method == 'POST':
        address.delete()
        return redirect('address_list')
    return render(request, 'address_confirm_delete.html', {'address': address})


@login_required
def cart(request):
    customer = request.user.customer
    cart_items = Cart.objects.filter(customer=customer)
    total_price = sum(item.product.price * item.quantity for item in cart_items)
    context = {
        'cart_items': cart_items,
        'total_price': total_price,
    }
    return render(request, 'cart.html', context)

@login_required(login_url='login')
def add_to_cart(request):
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        quantity = request.POST.get('quantity')
        if product_id and quantity:
            try:
                product = Product.objects.get(pk=product_id)
                customer = request.user.customer
                cart_item, created = Cart.objects.get_or_create(customer=customer, product=product)
                cart_item.quantity += int(quantity)
                if cart_item.quantity <= product.quantity_in_stock:
                    cart_item.save()
                    messages.success(request, f'{quantity} item(s) added to cart.')
                else:
                    messages.error(request, 'Requested quantity exceeds available stock.')
            except Product.DoesNotExist:
                messages.error(request, 'Product does not exist.')
        else:
            messages.error(request, 'Invalid request.')
    return redirect('cart')

@login_required
def delete_item_in_cart(request, id):
    customer = request.user.customer
    product = get_object_or_404(Product, id=id)
    cart_item = Cart.objects.get(customer=customer, product=product)
    cart_item.delete()
    return redirect('cart')


@login_required
def increase_quantity(request, cart_item_id):
    cart_item = Cart.objects.get(pk=cart_item_id)
    if cart_item.quantity < cart_item.product.quantity_in_stock:
        cart_item.quantity += 1
        cart_item.save()
    return redirect('cart')

@login_required
def decrease_quantity(request, cart_item_id):
    cart_item = Cart.objects.get(pk=cart_item_id)
    if cart_item.quantity > 1:
        cart_item.quantity -= 1
        cart_item.save()
    else:
        cart_item.delete()
    return redirect('cart')


@login_required
def checkout(request):
    customer = request.user.customer
    cart_items = Cart.objects.filter(customer=customer)
    total_price = sum(item.product.price * item.quantity for item in cart_items)
    
    if total_price == 0:
        return redirect('order_list')  # Redirect to order details with a placeholder order id
    
    if request.method == 'POST':
        address_id = request.POST.get('address_id')
        if not address_id:
            messages.error(request, "Please select or add an address to checkout.")
            return redirect('checkout')
        address = Address.objects.get(id=address_id)
        order = Order.objects.create(customer=customer, address=address)
        for item in cart_items:
            OrderItem.objects.create(order=order, product=item.product, quantity=item.quantity)
            # Reduce the quantity of the product in stock
            item.product.quantity_in_stock -= item.quantity
            item.product.save()
            item.delete()
        
        # Send confirmation email to customer
        email_subject = 'Order Confirmation'
        email_body_html = render_to_string('order_confirmation_email.html', {'order': order})
        email_body_text = "Thank you for your order. Your order ID is {}. We will process it shortly.".format(order.id)
        email = EmailMultiAlternatives(
            email_subject,
            email_body_text,
            settings.EMAIL_HOST_USER,
            [customer.email],
        )
        email.attach_alternative(email_body_html, 'text/html')
        email.send()
        
        # Generate PDF bill
        pdf_template = get_template('bill_template.html')
        html = pdf_template.render({'order': order})
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'filename="bill.pdf"'
        #pisa_status = pisa.CreatePDF(html, dest=response)
        #if pisa_status.err:
        #    return HttpResponse('We had some errors <pre>' + html + '</pre>')
        #return response
        
        return redirect('order_detail', order.id)
    else:
        addresses = Address.objects.filter(customer=customer)
        context = {
            'cart_items': cart_items,
            'total_price': total_price,
            'addresses': addresses,
        }
        return render(request, 'checkout.html', context)
@login_required
def order_list(request):
    customer = request.user.customer
    orders = Order.objects.filter(customer=customer)
    context = {
        'orders': orders,
    }
    return render(request, 'order_list.html', context)

@login_required
def order_detail(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    context = {
        'order': order,
    }
    return render(request, 'order_detail.html', context)

def category_products(request, category_id):
    category = get_object_or_404(Category, pk=category_id)
    products = Product.objects.filter(subcategory__parent_category=category)
    return render(request, 'category_products.html', {'category': category, 'products': products})

def subcategory_products(request, subcategory_id):
    subcategory = get_object_or_404(Subcategory, pk=subcategory_id)
    products = Product.objects.filter(subcategory=subcategory)
    return render(request, 'subcategory_products.html', {'subcategory': subcategory, 'products': products})

@login_required
def order_list(request):
    customer = request.user.customer
    orders = Order.objects.filter(customer=customer)
    context = {
        'orders': orders,
    }
    return render(request, 'order_list.html', context)

@login_required
def order_detail(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    context = {
        'order': order,
    }
    return render(request, 'order_detail.html', context)

class SellerLoginView(View):
    def get(self, request):
        return render(request, 'Seller_login.html')

    def post(self, request):
        username = request.POST['username']
        password = request.POST['password']
        seller = authenticate(request, username=username, password=password)

        if seller is not None:
            login(request, seller)
            return redirect('seller_purchase_orders')
        else:
            messages.error(request, 'Invalid email or password. Please try again.')
            return render(request, 'Seller_login.html')
        
class CreatePurchaseOrderView(View):
    def get(self, request):
        sellers = Seller.objects.all()
        products = Product.objects.none()  # Initially empty

        if 'seller' in request.GET:
            seller_id = request.GET.get('seller')
            if seller_id:
                products = Product.objects.filter(seller_id=seller_id)
        
        context = {
            'sellers': sellers,
            'products': products,
        }
        return render(request, 'create_purchase_order.html', context)
    
    def post(self, request):
        seller_id = request.POST.get('seller')
        total_amount = request.POST.get('total_amount')

        # Get the selected seller
        selected_seller = Seller.objects.get(id=seller_id)

        # Create the PurchaseOrder object with the selected seller
        purchase_order = PurchaseOrder.objects.create(
            TotalAmount= Decimal(total_amount),
            PurchaseOrderDate=date.today(),
            Seller=selected_seller,  # Assign the Seller object, not just the ID
        )

        # Save purchase order items
        for i in range(len(request.POST.getlist('product'))):
            product_id = request.POST.getlist('product')[i]
            product = Product.objects.get(id=product_id)
            quantity = request.POST.getlist('quantity')[i]
            purchase_unit_price = Product.objects.get(id=product_id).cost
            
            PurchaseOrderItem.objects.create(
                Product=product,
                Quantity=quantity,
                PurchaseUnitPrice=purchase_unit_price,
                PurchaseOrder=purchase_order,
                TotalAmount=Decimal(quantity) * purchase_unit_price
            )

        return redirect('/admin/beautystore_app/purchaseorder/')  # Redirect to a success page
    
def seller_purchase_orders(request):
    # Assuming you have a way to identify the current seller, e.g., request.user.seller
    seller = request.user.seller
    purchase_orders = PurchaseOrder.objects.filter(Seller=seller)
    return render(request, 'seller_purchase_orders.html', {'purchase_orders': purchase_orders})

from django.db import transaction

@transaction.atomic
def purchase_order_details(request, purchase_order_id):
    purchase_order = get_object_or_404(PurchaseOrder, id=purchase_order_id)
    order_items = PurchaseOrderItem.objects.filter(PurchaseOrder=purchase_order)
    if request.method == 'POST':
        # Handle form submission to update delivery date and status
        delivery_date = request.POST.get('delivery_date')
        status = request.POST.get('status')
        # Update purchase order with new delivery date and status
        purchase_order.ExpectedDeliveryDate = delivery_date
        purchase_order.Status = status
        purchase_order.save()

        # Update product quantity if status is "Delivered"
        if status == 'Delivered':
            for item in order_items:
                item.Product.quantity_in_stock += item.Quantity
                item.Product.save()
        return redirect('seller_purchase_orders')

    return render(request, 'purchase_order_details.html', {'purchase_order': purchase_order, 'order_items': order_items})

def reject_purchase_order(request, purchase_order_id):
    if request.method == 'GET':
        seller_message = request.GET.get('seller_message', '')  # Get seller message from the query parameters
        purchase_order = get_object_or_404(PurchaseOrder, id=purchase_order_id)  # Get the purchase order object

        # Update purchase order status and seller message
        purchase_order.Status = 'Rejected'
        purchase_order.seller_message = seller_message
        purchase_order.save()

        return redirect('seller_purchase_orders')  # Redirect to seller purchase orders page
    

    
from django.db.models import Sum, F
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime
from .models import Order, OrderItem

def monthly_sales_report(request):
    # Get the month from the query parameters, default to the current month and year
    month = request.GET.get('month', datetime.now().strftime('%Y-%m'))
    
    # Split the month and year for query
    year, month = map(int, month.split('-'))

    # Calculate the start and end dates for the selected month
    start_date = datetime(year, month, 1)
    if month < 12:
        end_date = datetime(year, month + 1, 1)
    else:
        end_date = datetime(year + 1, 1, 1)

    # Filter orders within the selected month
    orders = Order.objects.filter(order_date__range=[start_date, end_date])

    # Calculate total sales for the month
    total_sales = sum(order.total_price() for order in orders) or 0

    # Calculate sales per product for the month
    order_items = OrderItem.objects.filter(order__in=orders)
    product_sales = order_items.values('product__name').annotate(
        total_quantity=Sum('quantity'),
        total_price=Sum(F('quantity') * F('product__price'))
    ).order_by('-total_price')

    if 'export' in request.GET:
        # Create an Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Sales Report {month}-{year}".replace("/", "-")

        # Write the headers
        headers = ['Product', 'Quantity Sold', 'Total Sales Amount']
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header

        # Write data to the Excel sheet
        for row_num, item in enumerate(product_sales, 2):
            ws[f"A{row_num}"] = item['product__name']
            ws[f"B{row_num}"] = item['total_quantity']
            ws[f"C{row_num}"] = item['total_price']

        # Add a row for total sales
        ws[f"A{row_num + 1}"] = 'Total Sales'
        ws[f"C{row_num + 1}"] = total_sales

        # Prepare the response to download the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Sales_Report_{month}_{year}.xlsx'
        wb.save(response)
        return response

    # Context for rendering the HTML page
    context = {
        'total_sales': total_sales,
        'month': f"{year}-{month:02d}",
        'product_sales': product_sales,
    }
    return render(request, 'admin_report.html', context)


from django.shortcuts import render
import google.generativeai as genai

# Configure Gemini API with your API key
genai.configure(api_key="AIzaSyAEpxEaoSLL6Z6gBM3Ha0edMAECjW6h61g")

def book_recommendation(request):
    recommendation = None
    
    if request.method == "POST":
        author = request.POST['author']
        genre = request.POST['genre']
        rating = request.POST['rating']
        price = request.POST['price']
        
        # Define the prompt to send to the Gemini API
        user_input = f"""
        Author: {author}
        Genre: {genre}
        Rating: {rating}
        Price: {price}
        Recommend a book based on the above criteria and provide a description.
        """
        
        # Call the Gemini model
        model = genai.GenerativeModel(model_name="gemini-1.5-flash")
        chat_session = model.start_chat(history=[])
        response = chat_session.send_message(user_input)
        
        recommendation = response.text  # Fetch the recommendation result
    
    return render(request, 'book_recommendation.html', {'recommendation': recommendation})


import json
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Count, Avg
from django.shortcuts import render
from .models import Product

def visualization_view(request):
    # Query for top 5 authors based on the number of books
    top_authors_data = Product.objects.values('author').annotate(count=Count('author')).order_by('-count')[:5]

    # Query for genre distribution
    genre_data = Product.objects.values('genre').annotate(count=Count('genre'))

    # Query for price distribution (average price per book)
    price_data = Product.objects.values('published_date').annotate(avg_price=Avg('price')).order_by('published_date')

    # Convert QuerySets to JSON serializable lists
    top_authors_data = json.dumps(list(top_authors_data), cls=DjangoJSONEncoder)
    genre_data = json.dumps(list(genre_data), cls=DjangoJSONEncoder)
    price_data = json.dumps(list(price_data), cls=DjangoJSONEncoder)

    # Pass data to the template
    context = {
        'top_authors_data': top_authors_data,
        'genre_data': genre_data,
        'price_data': price_data,
    }

    return render(request, 'visualizations.html', context)

from django.shortcuts import render
from .models import Cart, Product
from django.contrib.auth.decorators import login_required

@login_required
def similar_products(request):
    customer = request.user.customer
    
    # Get the products in the user's cart
    cart_items = Cart.objects.filter(customer=customer)
    
    if not cart_items.exists():
        return render(request, 'similar_products.html', {'message': "No products found in your cart."})
    
    # Get the subcategories of the products in the cart
    cart_products = [item.product for item in cart_items]
    categories = {product.subcategory for product in cart_products}
    
    # Filter similar products in the same subcategories, excluding the products in the cart
    similar_products = Product.objects.filter(subcategory__in=categories).exclude(id__in=[product.id for product in cart_products])
    
    # Check if similar products exist
    message = "No similar products found." if not similar_products.exists() else ""
    
    context = {
        'cart_products': cart_products,
        'similar_products': similar_products,
        'message': message
    }
    
    return render(request, 'similar_products.html', context)


import requests
from django.http import JsonResponse
from django.shortcuts import render

# Replace with your API key
API_KEY = 'AIzaSyDWn7FhpBcLeXw4ygXZd-86c-es1LJiWBA'
API_URL = 'https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash-latest:generateContent'

# Define predefined responses
PREDEFINED_RESPONSES = {
    "Can you tell me what this site is about?": "Hi, I'm here to help you find the perfect book! Our site provides personalized book recommendations based on your preferences, including genre, author, rating, and more.",
    "What kinds of books do you recommend?": "We recommend a wide range of books across various genres like fiction, non-fiction, mystery, romance, science fiction, fantasy, and more! Let me know what you're in the mood for.",
    "Can you recommend a book for me?": "Of course! Tell me a bit about your favorite genre, author, or what you're in the mood to read, and I'll suggest something you'll love.",
    "Can you give me details about a specific book?": "I'd be happy to! Just provide the title or author of the book you're interested in, and I'll give you all the details.",
    "How do I search for a book by author?": "Simply enter the author's name in the search bar, and I'll show you a list of their books with ratings, prices, and descriptions.",
    "What genres do you have?": "We offer books in a variety of genres, including mystery, romance, thriller, fantasy, science fiction, historical fiction, and many more. Which genre are you interested in?",
    "Can you recommend books based on my favorite genre?": "Sure! Just tell me your favorite genre, and I’ll recommend a few popular books you might enjoy.",
    "How do I add a book to my wishlist?": "To add a book to your wishlist, click on the 'Add to Wishlist' button next to the book you're interested in.",
    "How do I remove a book from my wishlist?": "To remove a book, simply go to your wishlist, find the book you want to remove, and click 'Remove'.",
    "Can I search for books by rating or price?": "Yes, you can! Use our filters to search for books based on their ratings or price range, and I'll help you find the best options.",
    "How do I check out?": "Once you’ve selected your books, click on your cart and proceed to checkout. Follow the steps to enter your shipping and payment information.",
    "I have a problem with my order. Who should I contact?": "If you have any issues with your order, please reach out to our support team through the contact page, or send an email to our support team.",
    "How can I track my order?": "You can track your order by visiting the 'Order Tracking' section and entering your order number.",
    "How do I create an account?": "To create an account, click on 'Sign Up', fill in the required details, and submit the form. You’ll receive a confirmation email to complete the process.",
    "How can I reset my password?": "If you’ve forgotten your password, go to the 'Login' page and click 'Forgot Password.' Follow the steps to reset it.",
    "What are the available payment options?": "We accept major credit and debit cards, as well as PayPal and net banking. Choose your preferred payment option during checkout.",
    "How can I contact customer support?": "You can contact customer support through our contact form or email our team at booksupport@example.com.",
    "Tell me a joke.": "Why don't books go on dates? Because they have too many characters!",
    "What’s your favorite book?": "I don't have a favorite book since I love all kinds of stories, but I can help you find one that will become your favorite!",
    "How do I rate a book?": "You can rate a book by visiting its detail page and selecting the number of stars you'd like to give. Your feedback helps others too!",
    "What are some popular books right now?": "Some of the current popular books include bestsellers across various genres. Let me know what genre you're interested in, and I can recommend something hot!",
    "Can I find books by a specific publisher?": "Yes, you can search for books by publisher using our search filters. Enter the publisher's name, and I'll list all relevant books.",
    "How do I leave a review for a book?": "You can leave a review by visiting the book’s detail page and clicking on 'Write a Review'. Share your thoughts to help other readers!"
}



def chat(request):
    if request.method == 'POST':
        user_message = request.POST.get('message')

        # Retrieve or initialize conversation history
        conversation_history = request.session.get('conversation_history', [])

        # Add user message to conversation history
        conversation_history.append(f"input: {user_message}")

        # Check if the message matches any predefined response
        bot_reply = PREDEFINED_RESPONSES.get(user_message, None)
        
        if not bot_reply:
            # Define headers and data for the API request
            headers = {
                'Content-Type': 'application/json',
            }
            
            # Prepare context: Use the conversation history
            messages = [{'text': message} for message in conversation_history]
            
            # Prepare data with context (previous conversation)
            data = {
                'contents': [
                    {
                        'parts': messages
                    }
                ]
            }

            # Make the API request
            try:
                response = requests.post(f'{API_URL}?key={API_KEY}', headers=headers, json=data)
                response.raise_for_status()  # Raise an exception for HTTP errors
                
                # Parse the JSON response
                api_response = response.json()
                print("API Response:", api_response)  # For debugging
                
                # Extract the bot reply from the response
                bot_reply = api_response['candidates'][0]['content']['parts'][0]['text']
                
                # Limit the response to a certain number of sentences (e.g., 3)
                bot_reply = '. '.join(bot_reply.split('. ')[:3])  # Limits the response to 3 sentences
                
            except requests.RequestException as e:
                # Handle request errors
                print(f"API request error: {e}")
                bot_reply = 'Sorry, there was an error processing your request.'

        # Add bot response to conversation history
        conversation_history.append(f"output: {bot_reply}")

        # Store updated conversation history in session
        request.session['conversation_history'] = conversation_history

        return JsonResponse({'reply': bot_reply})

    # Render the chat interface if not a POST request
    return render(request, 'chatbot.html')

from django.db.models import Sum, F
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl.utils import get_column_letter
import openpyxl
from datetime import datetime
from .models import Customer, Product, Order, OrderItem

def customer_report(request):
    # Get all customers from the database
    customers = Customer.objects.all()

    if 'export' in request.GET:
        # Create an Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Report"

        # Write the headers
        headers = ['Name', 'Email', 'Contact Number', 'Role']
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header

        # Write data to the Excel sheet
        for row_num, customer in enumerate(customers, 2):
            ws[f"A{row_num}"] = customer.customer_name
            ws[f"B{row_num}"] = customer.email
            ws[f"C{row_num}"] = customer.contact_number
            ws[f"D{row_num}"] = customer.role

        # Prepare the response to download the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Customer_Report.xlsx'
        wb.save(response)
        return response

    # Context for rendering the HTML page
    context = {
        'customers': customers,
    }
    return render(request, 'customer_report.html', context)


def stock_report(request):
    # Get all products from the database
    products = Product.objects.all()

    if 'export' in request.GET:
        # Create an Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Report"

        # Write the headers
        headers = ['Product Name', 'Price', 'Stock', 'Reorder Level', 'Status', 'Brand', 'Description']
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header

        # Write data to the Excel sheet
        for row_num, product in enumerate(products, 2):
            ws[f"A{row_num}"] = product.name
            ws[f"B{row_num}"] = product.price
            ws[f"C{row_num}"] = product.quantity_in_stock
            ws[f"D{row_num}"] = product.reorder_level
            ws[f"E{row_num}"] = product.stock_status()
            ws[f"F{row_num}"] = product.brand
            ws[f"G{row_num}"] = product.description

        # Prepare the response to download the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Stock_Report.xlsx'
        wb.save(response)
        return response

    # Context for rendering the HTML page
    context = {
        'products': products,
    }
    return render(request, 'stock_report.html', context)